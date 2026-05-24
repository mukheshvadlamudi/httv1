import openpyxl
import json
import os

file_path = "Futurelab Knowledge Vault.xlsx"
out_dir = os.path.join("apps", "web", "src", "data")
out_file = os.path.join(out_dir, "vault-resources.json")

def main():
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return

    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path)
    
    # 1. Parse Websites sheet
    websites = []
    if "Websites" in wb.sheetnames:
        sheet = wb["Websites"]
        rows = list(sheet.iter_rows(values_only=True))
        for r in rows:
            if len(r) >= 6 and r[2] is not None and r[2] != 'Resource' and r[2] != 'MDN Web Docs' and isinstance(r[1], (int, float)):
                websites.append({
                    "id": int(r[1]),
                    "resource": str(r[2]).strip(),
                    "url": str(r[3]).strip() if r[3] else "",
                    "category": str(r[4]).strip() if r[4] else "General",
                    "why_useful": str(r[5]).strip() if r[5] else ""
                })
        # Add MDN Web Docs back specifically since it is row 1
        for r in rows:
            if len(r) >= 6 and r[2] == 'MDN Web Docs':
                websites.append({
                    "id": 1,
                    "resource": "MDN Web Docs",
                    "url": str(r[3]).strip() if r[3] else "",
                    "category": str(r[4]).strip() if r[4] else "Frontend / Web fundamentals",
                    "why_useful": str(r[5]).strip() if r[5] else ""
                })
                break

    # 2. Parse YouTube sheet
    youtube = []
    if "YouTube" in wb.sheetnames:
        sheet = wb["YouTube"]
        rows = list(sheet.iter_rows(values_only=True))
        for r in rows:
            # Row index matches for '#', 'YouTuber / Channel', 'YouTube Channel URL'
            # Let's search for headers or rows where column index 2 has content and column index 1 has number
            if len(r) >= 4 and r[2] is not None and r[2] != 'YouTuber / Channel' and isinstance(r[1], (int, float)):
                youtube.append({
                    "id": int(r[1]),
                    "channel": str(r[2]).strip(),
                    "url": str(r[3]).strip() if r[3] else ""
                })

    # Sort websites by ID
    websites.sort(key=lambda x: x["id"])
    
    output_data = {
        "websites": websites,
        "youtube": youtube
    }

    # Ensure output directory exists
    os.makedirs(out_dir, exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
        
    print(f"Successfully converted Excel sheets. Saved to {out_file}.")
    print(f"Parsed {len(websites)} website resources and {len(youtube)} YouTube channels.")

if __name__ == "__main__":
    main()
